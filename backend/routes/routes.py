from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from extensions import db, bcrypt
from models.models import *
import os, io, requests as req_lib
from pypdf import PdfWriter, PdfReader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

main = Blueprint('main', __name__)

# ==================== ROLE MAP ====================

ROLE_MAP = {
    'super-admin': 'Super Admin',
    'admin':       'Admin',
    'staff':       'Staff'
}

SEMESTER_PROGRAM_MAP = {
    'Sem 1': None,
    'Sem 2': 'FYMCA-SEM-II',   # Only Sem 2 has electives (310916A-E)
    'Sem 3': None,
    'Sem 4': None,
}

# ==================== EXCEL HELPERS ====================

def make_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_cell(cell, bg_color='1F4E79'):
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.fill      = PatternFill('solid', start_color=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = make_border()

def style_data_cell(cell, center=False):
    cell.font      = Font(name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center', wrap_text=True)
    cell.border    = make_border()

def style_link_cell(cell):
    cell.font      = Font(name='Arial', size=10, color='0563C1', underline='single')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border    = make_border()

# Dead UI routes removed. React handles /, /login, /register, and /dashboard.

# ==================== API ====================

@main.route('/api/search_courses')
def search_courses():
    query     = request.args.get('q', '')
    results   = Course.query.filter(
        (Course.course_name.ilike(f'%{query}%')) |
        (Course.course_code.ilike(f'%{query}%'))
    ).all()
    added_ids = {r.course_id for r in C11_Courses.query.all()}
    return jsonify([{
        'id': c.id, 'name': c.course_name, 'code': c.course_code,
        'year': c.year_of_intro, 'duplicate': c.id in added_ids
    } for c in results])

@main.route('/api/courses_by_program/<int:program_id>')
def courses_by_program(program_id):
    courses   = Course.query.filter_by(program_id=program_id).order_by(Course.course_code).all()
    added_ids = {r.course_id for r in C11_Courses.query.all()}
    return jsonify([{
        'id': c.id, 'name': c.course_name, 'code': c.course_code,
        'year': c.year_of_intro, 'duplicate': c.id in added_ids
    } for c in courses])

@main.route('/api/subjects_by_semester/<semester>')
def subjects_by_semester(semester):
    prog_name = SEMESTER_PROGRAM_MAP.get(semester)
    if not prog_name:
        return jsonify([])
    program = Program.query.filter_by(program_name=prog_name).first()
    if not program:
        program = Program.query.filter(Program.program_name.ilike('%SEM-II%')).first()
    if not program:
        return jsonify([])
    courses = Course.query.filter(
        Course.program_id == program.id,
        Course.course_code.like('310916%')
    ).order_by(Course.course_code).all()
    return jsonify([{
        'code': c.course_code,
        'name': c.course_name,
        'year': c.year_of_intro
    } for c in courses])

@main.route('/api/academic_years')
def get_academic_years():
    years = AcademicYear.query.order_by(AcademicYear.year_name).all()
    return jsonify([{'id': y.id, 'name': y.year_name} for y in years])

@main.route('/api/addon_subjects/<year_level>')
def get_addon_subjects(year_level):
    subjects = AddonSubject.query.filter_by(year_level=year_level).all()
    return jsonify([{'id': s.id, 'name': s.name} for s in subjects])

@main.route('/api/add_addon_subject', methods=['POST'])
def add_addon_subject():
    data = request.json
    name = data.get('name')
    year_level = data.get('year_level')
    if name and year_level:
        subj = AddonSubject(name=name, year_level=year_level)
        db.session.add(subj)
        db.session.commit()
        return jsonify({'success': True, 'id': subj.id, 'name': subj.name})
    return jsonify({'success': False}), 400

@main.route('/api/students', methods=['GET'])
def get_students():
    students = Student.query.all()
    return jsonify([{
        'id': s.id, 
        'name': s.name, 
        'enrollment_number': s.enrollment_number, 
        'program_code': s.program_code,
        'category': s.category
    } for s in students])

@main.route('/api/students', methods=['POST'])
def add_student():
    data = request.json
    name = data.get('name')
    enrollment_number = data.get('enrollment_number')
    program_code = data.get('program_code')
    category = data.get('category')
    
    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
        
    if enrollment_number:
        existing = Student.query.filter_by(enrollment_number=enrollment_number).first()
        if existing:
            return jsonify({'success': False, 'message': 'Student with this enrollment number already exists'}), 400
            
    try:
        new_student = Student(
            name=name,
            enrollment_number=enrollment_number,
            program_code=program_code,
            category=category
        )
        db.session.add(new_student)
        db.session.commit()
        return jsonify({
            'success': True, 
            'student': {
                'id': new_student.id, 
                'name': new_student.name, 
                'enrollment_number': new_student.enrollment_number, 
                'program_code': new_student.program_code,
                'category': new_student.category
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/api/students/upload', methods=['POST'])
def upload_students():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400
        
    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            students_added = 0
            
            # Extract headers from first row
            headers = [cell.value for cell in ws[1]]
            name_idx = next((i for i, h in enumerate(headers) if h and 'name' in str(h).lower()), 0)
            enroll_idx = next((i for i, h in enumerate(headers) if h and 'enroll' in str(h).lower() or h and 'prn' in str(h).lower()), -1)
            prog_idx = next((i for i, h in enumerate(headers) if h and 'program' in str(h).lower()), -1)
            cat_idx = next((i for i, h in enumerate(headers) if h and 'category' in str(h).lower()), -1)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[name_idx] if name_idx >= 0 and name_idx < len(row) else None
                if not name or not str(name).strip() or str(name).strip().lower() == 'none':
                    continue
                
                enrollment = str(row[enroll_idx]).strip() if enroll_idx >= 0 and enroll_idx < len(row) and row[enroll_idx] is not None else None
                prog = str(row[prog_idx]).strip() if prog_idx >= 0 and prog_idx < len(row) and row[prog_idx] is not None else None
                category = str(row[cat_idx]).strip() if cat_idx >= 0 and cat_idx < len(row) and row[cat_idx] is not None else None
                
                if enrollment and enrollment.lower() != 'none':
                    existing = Student.query.filter_by(enrollment_number=enrollment).first()
                    if existing:
                        continue
                
                db.session.add(Student(
                    name=str(name).strip(),
                    enrollment_number=enrollment if enrollment and enrollment.lower() != 'none' else None,
                    program_code=prog if prog and prog.lower() != 'none' else None,
                    category=category if category and category.lower() != 'none' else None
                ))
                students_added += 1
                
            db.session.commit()
            return jsonify({'success': True, 'added': students_added, 'message': f'Successfully added {students_added} students'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        return jsonify({'success': False, 'message': 'Please upload a valid .xlsx file'}), 400

# ==================== CRITERIA 1.1 ====================

@main.route('/criteria/1.1')
def criterion1_1():
    if 'user_id' not in session:
        return redirect(url_for('main.home'))
    data          = C11_Courses.query.all()
    programs_list = Program.query.all()
    proof_links   = ProofLink11.query.first()
    return render_template('criteria1.html', data=data,
                           programs_list=programs_list, proof_links=proof_links)

@main.route('/add_record/1.1', methods=['POST'])
def add_record_11():
    if 'user_id' not in session:
        return redirect(url_for('main.home'))
    course_ids = request.form.getlist('course_id[]')
    program_id = request.form.get('program_id')
    year       = request.form.get('year')
    dept_id    = None
    if program_id:
        prog = Program.query.get(program_id)
        if prog:
            dept_id = prog.dept_id
    added_ids = {r.course_id for r in C11_Courses.query.all()}
    skipped = saved = 0
    for c_id in course_ids:
        if not c_id: continue
        if int(c_id) in added_ids: skipped += 1; continue
        db.session.add(C11_Courses(academic_year=year,
                                  dept_id=dept_id, course_id=int(c_id)))
        saved += 1
    db.session.commit()
    if saved:   flash(f"{saved} record(s) saved successfully!", "success")
    if skipped: flash(f"{skipped} course(s) skipped — already exist.", "warning")
    if not saved and not skipped: flash("No courses selected.", "warning")
    return redirect(url_for('main.criterion1_1'))

@main.route('/delete_record/<int:id>')
def delete_record(id):
    rec = C11_Courses.query.get_or_404(id)
    db.session.delete(rec); db.session.commit()
    flash("Record deleted.", "success")
    return redirect(url_for('main.criterion1_1'))

@main.route('/upload_proof_11', methods=['POST'])
def upload_proof_11():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    link1 = request.form.get('link1', '').strip()
    link2 = request.form.get('link2', '').strip()
    saved = ProofLink11.query.first()
    if saved:
        saved.link1 = link1; saved.link2 = link2
    else:
        db.session.add(ProofLink11(link1=link1, link2=link2))
    db.session.commit()
    flash("Proof links saved successfully!", "success")
    return redirect(url_for('main.criterion1_1'))

@main.route('/export/11/excel')
def export_11_excel():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records     = C11_Courses.query.order_by(C11_Courses.id).all()
    proof_links = ProofLink11.query.first()
    link1 = proof_links.link1 if proof_links else ''
    link2 = proof_links.link2 if proof_links else ''

    wb = Workbook(); ws = wb.active; ws.title = 'Criteria 1.1'
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value     = '1.1 Number of courses offered by the Institution across all programs during the year'
    t.font      = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    t.fill      = PatternFill('solid', start_color='C00000')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value     = 'MET Bhujbal Knowledge City – Institute of Engineering, Nashik'
    s.font      = Font(bold=True, size=10, name='Arial')
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    headers = ['#', 'Program Name', 'Course Code', 'Course Name',
               'Year of Introduction', 'Academic Year', 'Proof Link 1', 'Proof Link 2']
    for col, h in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=3, column=col, value=h))
    ws.row_dimensions[3].height = 30

    row_num = 4
    for i, rec in enumerate(records):
        program_name  = rec.course.program.program_name if rec.course and rec.course.program else '—'
        course_code   = rec.course.course_code   if rec.course else '—'
        course_name   = rec.course.course_name   if rec.course else '—'
        year_of_intro = rec.course.year_of_intro if rec.course else 2020
        academic_year = rec.academic_year or '—'
        style_data_cell(ws.cell(row=row_num, column=1, value=i + 1), center=True)
        style_data_cell(ws.cell(row=row_num, column=2, value=program_name), center=True)
        style_data_cell(ws.cell(row=row_num, column=3, value=course_code), center=True)
        style_data_cell(ws.cell(row=row_num, column=4, value=course_name))
        style_data_cell(ws.cell(row=row_num, column=5, value=year_of_intro), center=True)
        style_data_cell(ws.cell(row=row_num, column=6, value=academic_year), center=True)
        c7 = ws.cell(row=row_num, column=7, value=link1 if link1 else '—')
        style_link_cell(c7) if link1 else style_data_cell(c7, center=True)
        c8 = ws.cell(row=row_num, column=8, value=link2 if link2 else '—')
        style_link_cell(c8) if link2 else style_data_cell(c8, center=True)
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    if records:
        merge_start = 4
        prev_prog = records[0].course.program.name if records[0].course and records[0].course.program else '—'
        for idx in range(1, len(records)):
            prog_name = records[idx].course.program.name if records[idx].course and records[idx].course.program else '—'
            data_row  = 4 + idx
            if prog_name != prev_prog:
                if data_row - 1 > merge_start:
                    ws.merge_cells(f'B{merge_start}:B{data_row - 1}')
                    ws[f'B{merge_start}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                merge_start = data_row; prev_prog = prog_name
        last_row = 3 + len(records)
        if last_row > merge_start:
            ws.merge_cells(f'B{merge_start}:B{last_row}')
            ws[f'B{merge_start}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if len(records) > 1:
        last_data_row = 3 + len(records)
        ws.merge_cells(f'G4:G{last_data_row}')
        ws['G4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(f'H4:H{last_data_row}')
        ws['H4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for i, w in enumerate([5, 18, 14, 46, 20, 14, 45, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, download_name='Criteria_1_1_Courses.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== CRITERIA 1.2.1 ====================

@main.route('/criteria/1.2.1')
def criterion1_2_1():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    return render_template('criteria1.2.1.html', data=C121_CBCS.query.all(),
                           saved_links=ProofLink121.query.first())

@main.route('/add_121', methods=['POST'])
def add_121():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    semester      = request.form.get('semester')
    prog_codes    = request.form.getlist('prog_code[]')
    prog_names    = request.form.getlist('prog_name[]')
    year_intros   = request.form.getlist('year_intro[]')
    cbcs_statuses = request.form.getlist('cbcs_status[]')
    cbcs_years    = request.form.getlist('cbcs_year[]')
    for i in range(len(prog_codes)):
        if prog_codes[i]:
            existing = C121_CBCS.query.filter_by(program_code=prog_codes[i]).first()
            if existing:
                flash(f"Subject Code {prog_codes[i]} is already added. Skipped duplicate.", "warning")
                continue
                
            db.session.add(C121_CBCS(
                program_code=prog_codes[i],
                program_name=prog_names[i]    if i < len(prog_names)    else '',
                year_of_introduction=year_intros[i]  if i < len(year_intros)   else '',
                cbcs_status=cbcs_statuses[i] if i < len(cbcs_statuses) else 'Yes',
                cbcs_year=cbcs_years[i]    if i < len(cbcs_years)    else ''))
    db.session.commit()
    flash("Records saved successfully!", "success")
    return redirect(url_for('main.criterion1_2_1'))

@main.route('/edit_121/<int:id>', methods=['POST'])
def edit_121(id):
    rec = C121_CBCS.query.get_or_404(id)
    rec.program_code   = request.form.get('prog_code')
    rec.program_name   = request.form.get('prog_name')
    rec.year_of_introduction  = request.form.get('year_intro')
    rec.cbcs_status = request.form.get('cbcs_status')
    rec.cbcs_year   = request.form.get('cbcs_year')
    db.session.commit(); flash("Record updated!", "success")
    return redirect(url_for('main.criterion1_2_1'))

@main.route('/delete_121/<int:id>')
def delete_121(id):
    rec = C121_CBCS.query.get_or_404(id)
    db.session.delete(rec); db.session.commit()
    flash("Record deleted.", "success")
    return redirect(url_for('main.criterion1_2_1'))

@main.route('/upload_proof_121', methods=['POST'])
def upload_proof_121():
    syllabus_link = request.form.get('syllabus_link', '')
    elective_link = request.form.get('elective_link', '')
    saved = ProofLink121.query.first()
    if saved:
        saved.syllabus = syllabus_link; saved.elective = elective_link
    else:
        db.session.add(ProofLink121(syllabus=syllabus_link, elective=elective_link))
    db.session.commit(); flash("Proof links updated!", "success")
    return redirect(url_for('main.criterion1_2_1'))

@main.route('/export_121_excel')
def export_121_excel():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records     = C121_CBCS.query.all()
    proof_links = ProofLink121.query.first()
    link1 = (proof_links.syllabus or '').strip() if proof_links else ''
    link2 = (proof_links.elective or '').strip() if proof_links else ''

    wb = Workbook(); ws = wb.active; ws.title = 'Criteria 1.2.1'
    ws.merge_cells('A1:G1')
    h = ws['A1']
    h.value     = '1.2.1 Number of Programmes in which Choice Based Credit System (CBCS)/ elective course system has been implemented'
    h.font      = Font(bold=True, size=11, name='Arial')
    h.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    h.border    = make_border()
    ws.row_dimensions[1].height = 30

    headers = ['Programme Code', 'Programme name', 'Year of Introduction',
               'Status of implemetation of\nCBCS / elective course\nsystem (Yes/No)',
               'Year of implemetation of\nCBCS / elective course\nsystem',
               'Proof Link 1', 'Proof Link 2']
    for col, h in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=2, column=col, value=h))
    ws.row_dimensions[2].height = 50

    if records:
        style_data_cell(ws.cell(row=3, column=1, value='515124110'), center=True)
        style_data_cell(ws.cell(row=3, column=2, value='MCA'))
        for col in range(3, 8): style_data_cell(ws.cell(row=3, column=col))
        ws.row_dimensions[3].height = 18

        row_num = 4
        for rec in records:
            style_data_cell(ws.cell(row=row_num, column=1, value=rec.program_code), center=True)
            style_data_cell(ws.cell(row=row_num, column=2, value=rec.program_name))
            style_data_cell(ws.cell(row=row_num, column=3, value=rec.year_of_introduction))
            style_data_cell(ws.cell(row=row_num, column=4, value=rec.cbcs_status), center=True)
            style_data_cell(ws.cell(row=row_num, column=5, value=rec.cbcs_year), center=True)
            style_data_cell(ws.cell(row=row_num, column=6))
            style_data_cell(ws.cell(row=row_num, column=7))
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        first_data = 4; last_data = 3 + len(records)
        for col_letter, link in [('F', link1), ('G', link2)]:
            if last_data > first_data:
                ws.merge_cells(f'{col_letter}{first_data}:{col_letter}{last_data}')
            c = ws[f'{col_letter}{first_data}']
            c.value     = link if link else '—'
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border    = make_border()
            c.font      = Font(name='Arial', size=10,
                               color='0563C1' if link else '000000',
                               underline='single' if link else None)

        for _ in range(5):
            for col in range(1, 8): style_data_cell(ws.cell(row=row_num, column=col))
            ws.row_dimensions[row_num].height = 18; row_num += 1

    for i, w in enumerate([16, 30, 20, 30, 24, 45, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, download_name='Criteria_1_2_1_CBCS.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== CRITERIA 1.1.3 ====================

@main.route('/criteria/1.1.3')
def criterion1_1_3():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    return render_template('criteria1.1.3.html', data=C113_TeacherBodies.query.all())

@main.route('/add_113', methods=['POST'])
def add_113():
    years       = request.form.getlist('year[]')
    teachers    = request.form.getlist('teacher_name[]')
    bodies      = request.form.getlist('body_name[]')
    drive_links = request.form.getlist('drive_link[]')
    for i in range(len(years)):
        db.session.add(C113_TeacherBodies(
            year=years[i], teacher_name=teachers[i], body_name=bodies[i],
            drive_link=drive_links[i] if i < len(drive_links) else ''))
    db.session.commit(); flash("Records saved!", "success")
    return redirect(url_for('main.criterion1_1_3'))

@main.route('/edit_113/<int:id>', methods=['POST'])
def edit_113(id):
    rec              = C113_TeacherBodies.query.get_or_404(id)
    rec.year         = request.form.get('year')
    rec.teacher_name = request.form.get('teacher_name')
    rec.body_name    = request.form.get('body_name')
    rec.drive_link   = request.form.get('drive_link', '').strip()
    db.session.commit(); flash("Record updated!", "success")
    return redirect(url_for('main.criterion1_1_3'))

@main.route('/delete_113/<int:id>')
def delete_113(id):
    rec = C113_TeacherBodies.query.get_or_404(id)
    db.session.delete(rec); db.session.commit()
    flash("Record deleted.", "success")
    return redirect(url_for('main.criterion1_1_3'))

@main.route('/export_113_excel')
def export_113_excel():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records = C113_TeacherBodies.query.order_by(C113_TeacherBodies.id).all()

    wb = Workbook(); ws = wb.active; ws.title = 'Criteria 1.1.3'
    ws.merge_cells('A1:D1')
    h = ws['A1']
    h.value = (
        "1.1.3 Teachers of the Institution participate in following activities related to "
        "curriculum development and assessment of the affiliating University and/are "
        "represented on the following academic bodies during the year\n"
        "1. Academic council/BoS of Affiliating university  "
        "2. Setting of question papers for UG/PG programs  "
        "3. Design and Development of Curriculum for Add on/ certificate/ Diploma Courses  "
        "4. Assessment/evaluation process of the affiliating University"
    )
    h.font = Font(bold=False, size=10, name='Arial')
    h.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    h.border = make_border()
    ws.row_dimensions[1].height = 80
    ws.row_dimensions[2].height = 6

    headers = ['Year', 'Name of Teacher Participated',
               'Name of the Body in which Teacher Participated',
               'Proof Document (Drive Link)']
    for col, h in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=3, column=col, value=h))
    ws.row_dimensions[3].height = 36

    row_num = 4
    for rec in records:
        style_data_cell(ws.cell(row=row_num, column=1, value=rec.year), center=True)
        style_data_cell(ws.cell(row=row_num, column=2, value=rec.teacher_name))
        style_data_cell(ws.cell(row=row_num, column=3, value=rec.body_name))
        dl = rec.drive_link or ''
        c4 = ws.cell(row=row_num, column=4, value=dl if dl else '—')
        style_link_cell(c4) if dl else style_data_cell(c4, center=True)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    for _ in range(5):
        for col in range(1, 5): style_data_cell(ws.cell(row=row_num, column=col))
        ws.row_dimensions[row_num].height = 18; row_num += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.freeze_panes = 'A4'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, download_name='Criteria_1_1_3_Teachers.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main.route('/merge_113_pdfs')
def merge_113_pdfs():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records = C113_TeacherBodies.query.order_by(C113_TeacherBodies.id).all()
    links   = [(r.teacher_name, r.body_name, r.drive_link)
               for r in records if r.drive_link and r.drive_link.strip()]
    if not links:
        flash("No Google Drive links found.", "warning")
        return redirect(url_for('main.criterion1_1_3'))

    writer  = PdfWriter(); skipped = []; merged = 0
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    for teacher, body, link in links:
        link = link.strip(); file_id = None
        if '/file/d/' in link:
            try: file_id = link.split('/file/d/')[1].split('/')[0]
            except: pass
        elif 'id=' in link:
            try: file_id = link.split('id=')[1].split('&')[0]
            except: pass
        if not file_id:
            skipped.append(f"{teacher} (invalid link)"); continue
        try:
            sess = req_lib.Session()
            resp = sess.get(f'https://drive.google.com/uc?export=download&id={file_id}&confirm=t',
                            headers=headers, timeout=30, allow_redirects=True)
            if 'text/html' in resp.headers.get('Content-Type', ''):
                token = None
                for line in resp.text.split('\n'):
                    if 'confirm=' in line and 'export=download' in line:
                        try: token = line.split('confirm=')[1].split('&')[0].split('"')[0]
                        except: pass
                        break
                if token:
                    resp = sess.get(f'https://drive.google.com/uc?export=download&id={file_id}&confirm={token}',
                                    headers=headers, timeout=30)
            content = resp.content
            if not content.startswith(b'%PDF'):
                skipped.append(f"{teacher} (not a valid PDF)"); continue
            for page in PdfReader(io.BytesIO(content)).pages:
                writer.add_page(page)
            merged += 1
        except Exception as e:
            skipped.append(f"{teacher} ({str(e)[:60]})"); continue

    if merged == 0:
        flash("Could not download any PDFs. Make sure Drive links are public.", "danger")
        return redirect(url_for('main.criterion1_1_3'))

    output = io.BytesIO(); writer.write(output); output.seek(0)
    if skipped:
        flash(f"Combined {merged} PDF(s). Skipped: {', '.join(skipped)}", "warning")
    return send_file(output, download_name='Criteria_1_1_3_Combined_Proofs.pdf',
                     as_attachment=True, mimetype='application/pdf')

# ==================== CRITERIA 1.2.2 & 1.2.3 ====================

@main.route('/criteria/1.2.2_1.2.3')
def criterion1_2_2_1_2_3():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    return render_template('criteria1.2.2.html', data=C122_Addon.query.all(),
                           saved_links=ProofLink122_123.query.first())

@main.route('/add_122_123', methods=['POST'])
def add_122_123():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    year_level = request.form.get('year_level')
    academic_years = request.form.getlist('academic_year[]')
    program_names = request.form.getlist('program_name[]')
    course_codes = request.form.getlist('course_code[]')
    times_offered = request.form.getlist('times_offered[]')
    durations = request.form.getlist('duration[]')
    students_enrolled = request.form.getlist('students_enrolled[]')
    students_completed = request.form.getlist('students_completed[]')

    for i in range(len(program_names)):
        if program_names[i]:
            db.session.add(C122_Addon(
                year_level=year_level,
                academic_year=academic_years[i] if i < len(academic_years) else '',
                program_name=program_names[i],
                course_code=course_codes[i] if i < len(course_codes) else '',
                times_offered=int(times_offered[i]) if i < len(times_offered) and times_offered[i] else 0,
                duration=durations[i] if i < len(durations) else '',
                students_enrolled=int(students_enrolled[i]) if i < len(students_enrolled) and students_enrolled[i] else 0,
                students_completed=int(students_completed[i]) if i < len(students_completed) and students_completed[i] else 0
            ))
    db.session.commit()
    flash("Records saved successfully!", "success")
    return redirect(url_for('main.criterion1_2_2_1_2_3'))

@main.route('/edit_122_123/<int:id>', methods=['POST'])
def edit_122_123(id):
    rec = C122_Addon.query.get_or_404(id)
    rec.year_level = request.form.get('year_level')
    rec.academic_year = request.form.get('academic_year')
    rec.program_name = request.form.get('program_name')
    rec.course_code = request.form.get('course_code')
    rec.times_offered = int(request.form.get('times_offered') or 0)
    rec.duration = request.form.get('duration')
    rec.students_enrolled = int(request.form.get('students_enrolled') or 0)
    rec.students_completed = int(request.form.get('students_completed') or 0)
    db.session.commit()
    flash("Record updated!", "success")
    return redirect(url_for('main.criterion1_2_2_1_2_3'))

@main.route('/delete_122_123/<int:id>')
def delete_122_123(id):
    rec = C122_Addon.query.get_or_404(id)
    db.session.delete(rec)
    db.session.commit()
    flash("Record deleted.", "success")
    return redirect(url_for('main.criterion1_2_2_1_2_3'))

@main.route('/upload_proof_122_123', methods=['POST'])
def upload_proof_122_123():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    proof_links = request.form.getlist('proof_links[]')
    valid_links = [link.strip() for link in proof_links if link.strip()]
    
    saved = ProofLink122_123.query.first()
    if saved:
        saved.links = valid_links
    else:
        db.session.add(ProofLink122_123(links=valid_links))
    db.session.commit()
    flash("Proof links updated!", "success")
    return redirect(url_for('main.criterion1_2_2_1_2_3'))

@main.route('/export_122_123_excel')
def export_122_123_excel():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records = C122_Addon.query.order_by(C122_Addon.year_level, C122_Addon.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Criteria 1.2.2 & 1.2.3'
    
    ws.merge_cells('A1:G1')
    h = ws['A1']
    h.value = '1.2.2 Number of Add on /Certificate programs offered during the year\n1.2.3 Number of students enrolled in Certificate/ Add-on programs as against the total number of students during the year'
    h.font = Font(bold=True, size=11, name='Arial')
    h.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    h.border = make_border()
    ws.row_dimensions[1].height = 40

    headers = [
        'Name of Add on /Certificate programs offered', 
        'Course Code (if any)', 
        'Year of offering',
        'No. of times offered during the same year',
        'Duration of course',
        'Number of students enrolled in the year',
        'Students completing the course in the year'
    ]

    year_groups = {}
    for r in records:
        key = r.year_level or 'Unspecified Year'
        if key not in year_groups: year_groups[key] = []
        year_groups[key].append(r)

    sorted_years = sorted(year_groups.keys())

    row_num = 2
    for y_level in sorted_years:
        ay_example = year_groups[y_level][0].academic_year if year_groups[y_level] else ''
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        yh = ws.cell(row=row_num, column=1)
        yh.value = f"{y_level} (Academic Year {ay_example})" if ay_example else y_level
        yh.font = Font(bold=True, size=11, name='Arial')
        yh.alignment = Alignment(horizontal='center', vertical='center')
        yh.fill = PatternFill('solid', start_color='FFC000')
        yh.border = make_border()
        row_num += 1

        for col, head in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col, value=head)
            cell.font = Font(bold=True, size=10, name='Arial')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = make_border()
        ws.row_dimensions[row_num].height = 45
        row_num += 1

        for rec in year_groups[y_level]:
            style_data_cell(ws.cell(row=row_num, column=1, value=rec.program_name))
            style_data_cell(ws.cell(row=row_num, column=2, value=rec.course_code), center=True)
            style_data_cell(ws.cell(row=row_num, column=3, value=rec.academic_year), center=True)
            style_data_cell(ws.cell(row=row_num, column=4, value=rec.times_offered), center=True)
            style_data_cell(ws.cell(row=row_num, column=5, value=rec.duration), center=True)
            style_data_cell(ws.cell(row=row_num, column=6, value=rec.students_enrolled), center=True)
            style_data_cell(ws.cell(row=row_num, column=7, value=rec.students_completed), center=True)
            row_num += 1
        
        row_num += 2

    for i, w in enumerate([40, 20, 20, 20, 20, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='Criteria_1_2_2_1_2_3.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== CRITERIA 1.3.2 ====================

@main.route('/criteria/1.3.2')
def criterion1_3_2():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records = C132_Experiential.query.all()
    grouped = {}
    for r in records:
        key = (r.program_code, r.course_code, r.year_of_offering)
        if key not in grouped:
            grouped[key] = {
                'program_name': r.program_name,
                'program_code': r.program_code,
                'experiential_course_name': r.experiential_course_name,
                'course_code': r.course_code,
                'year_of_offering': r.year_of_offering,
                'student_count': 0
            }
        grouped[key]['student_count'] += 1

    grouped_data = list(grouped.values())
    return render_template('criteria1.3.2.html', grouped_data=grouped_data)

@main.route('/add_132', methods=['POST'])
def add_132():
    if 'user_id' not in session: return redirect(url_for('main.home'))
    
    program_name = request.form.get('program_name')
    program_code = request.form.get('program_code')
    course_name = request.form.get('experiential_course_name')
    course_code = request.form.get('course_code')
    year_of_offering = request.form.get('year_of_offering')
    proof_link = request.form.get('proof_link', '').strip()
    proof_links_json = [proof_link] if proof_link else []
    
    file = request.files.get('student_excel')
    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            students_added = 0
            # Iterate through rows starting from 2 (assuming header)
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                student_name = row[0]
                if student_name and str(student_name).strip():
                    db.session.add(C132_Experiential(
                        program_name=program_name,
                        program_code=program_code,
                        experiential_course_name=course_name,
                        course_code=course_code,
                        year_of_offering=year_of_offering,
                        student_name=str(student_name).strip(),
                        proof_links=proof_links_json
                    ))
                    students_added += 1
            db.session.commit()
            flash(f"Successfully processed Excel and added {students_added} students!", "success")
        except Exception as e:
            flash(f"Error processing Excel file: {str(e)}", "danger")
    else:
        flash("Please upload a valid .xlsx file.", "danger")

    return redirect(url_for('main.criterion1_3_2'))

@main.route('/delete_132/<prog_code>/<course_code>/<year>')
def delete_132(prog_code, course_code, year):
    if 'user_id' not in session: return redirect(url_for('main.home'))
    records = C132_Experiential.query.filter_by(
        program_code=prog_code,
        course_code=course_code,
        year_of_offering=year
    ).all()
    
    for r in records:
        db.session.delete(r)
    db.session.commit()
    flash(f"Deleted course {course_code} data.", "success")
    return redirect(url_for('main.criterion1_3_2'))


