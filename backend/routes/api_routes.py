import os
import uuid
from flask import Blueprint, jsonify, request, session, current_app, send_file
from werkzeug.utils import secure_filename
from models.models import *
from extensions import db, bcrypt
import pandas as pd
from datetime import datetime, date
from decimal import Decimal

api_bp = Blueprint('api_bp', __name__, url_prefix='/api')

# --- Auth Endpoints ---

@api_bp.route('/register', methods=['POST'])
def api_register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'Admin')
    department = data.get('department', '')
    user_identifier = data.get('user_identifier', '')
    year = data.get('year', '')

    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({"success": False, "error": "Username already exists"}), 400

    hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
    new_user = User(
        username=username, 
        password=hashed_pw, 
        role=role, 
        department=department,
        user_identifier=user_identifier,
        year=year
    )
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({"success": True, "message": "Account created"})


@api_bp.route('/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    academic_year = data.get('academic_year')
    
    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400
        
    user = User.query.filter_by(username=username).first()
    if user and bcrypt.check_password_hash(user.password, password):
        session['user_id'] = user.id
        session['academic_year'] = academic_year
        
        return jsonify({
            "success": True, 
            "user": {
                "id": user.id, 
                "username": user.username, 
                "role": user.role, 
                "department": user.department,
                "academic_year": academic_year or user.year,
                "program": "MCA",
                "programCode": "515124110"
            }
        })
    return jsonify({"success": False, "error": "Invalid username or password"}), 401


# --- Metadata Endpoints ---

@api_bp.route('/teachers', methods=['GET'])
def get_teachers():
    teachers = Teacher.query.all()
    return jsonify([{"id": t.id, "name": t.name} for t in teachers])

@api_bp.route('/departments', methods=['GET'])
def get_departments():
    progs = Program.query.all()
    res, seen = [], set()
    for p in progs:
        dept = p.department or p.program_name
        if dept not in seen:
            seen.add(dept)
            res.append({"id": p.id, "code": dept, "programCode": p.program_code, "programName": p.program_name})
    if not res:
        res = [{"id": 1, "code": "MCA", "programCode": "515124110", "programName": "MCA"}]
    return jsonify(res)

@api_bp.route('/semesters', methods=['GET'])
def get_semesters():
    recs = SemesterLookup.query.all()
    if not recs:
        # Default seed
        defaults = ["FYMCA-SEM-I (MCA)", "FYMCA-SEM-II (MCA)", "SYMCA-SEM-III (MCA)", "SYMCA-SEM-IV (MCA)"]
        for d in defaults: db.session.add(SemesterLookup(value=d))
        db.session.commit()
        recs = SemesterLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/courses', methods=['GET'])
def get_courses():
    """
    GET /api/courses          → all courses in SemesterCourseLookup
    GET /api/courses?sem=X    → courses for that semester only
    """
    sem = request.args.get('sem')
    q = SemesterCourseLookup.query
    if sem:
        q = q.filter(SemesterCourseLookup.semester == sem)
    rows = q.order_by(SemesterCourseLookup.course_code).all()
    return jsonify([{"code": r.course_code, "name": r.course_name, "semester": r.semester} for r in rows])

@api_bp.route('/courses', methods=['POST'])
def add_course():
    """
    POST /api/courses  { semester, course_code, course_name }
    Adds a new course to SemesterCourseLookup with normalisation + uniqueness check.
    """
    data = request.json or {}
    sem    = (data.get('semester') or '').strip()
    c_code = (data.get('course_code') or '').strip().upper()
    c_name = (data.get('course_name') or '').strip()

    if not sem or not c_code or not c_name:
        return jsonify({"success": False, "error": "semester, course_code and course_name are required"}), 400

    existing = SemesterCourseLookup.query.filter_by(semester=sem, course_code=c_code).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {c_code} already exists in {sem}"}), 400

    try:
        new_c = SemesterCourseLookup(semester=sem, course_code=c_code, course_name=c_name)
        db.session.add(new_c)
        db.session.commit()
        return jsonify({"success": True, "course": {"code": c_code, "name": c_name, "semester": sem}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


@api_bp.route('/electives', methods=['GET'])
def get_electives():
    # Attempting to load by 310916% or elective
    courses = Course.query.filter((Course.course_code.like('310916%')) | (Course.course_name.ilike('%elective%'))).all()
    return jsonify([{"id": c.id, "code": c.course_code, "name": c.course_name} for c in courses])

@api_bp.route('/program-codes', methods=['GET'])
def get_program_codes():
    progs = Program.query.all()
    return jsonify([{"code": p.program_code, "label": p.program_code} for p in progs])

@api_bp.route('/course-types', methods=['GET'])
def get_course_types():
    recs = CourseTypeLookup.query.all()
    if not recs:
        defaults = ["PBL (Project Based Learning)", "Major Project"]
        for d in defaults: db.session.add(CourseTypeLookup(value=d))
        db.session.commit()
        recs = CourseTypeLookup.query.all()
    # Note: legacy code returned courseCode. Keeping the structure.
    return jsonify([{"value": x.value, "label": x.value, "courseCode": ""} for x in recs])

@api_bp.route('/academic-years', methods=['GET'])
def get_academic_years():
    recs = AcademicYearLookup.query.all()
    if not recs:
        defaults = ["2020-21", "2021-22", "2022-23", "2023-24", "2024-25", "2025-26", "2026-27"]
        for d in defaults: db.session.add(AcademicYearLookup(value=d))
        db.session.commit()
        recs = AcademicYearLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/programmes', methods=['GET'])
def get_programmes():
    progs = Program.query.all()
    if not progs:
        return jsonify([
            {"code": "MCA",  "name": "Master of Computer Applications (MCA)"},
            {"code": "MBA",  "name": "Master of Business Administration (MBA)"}
        ])
    return jsonify([{"code": p.program_code, "name": p.program_name} for p in progs])

@api_bp.route('/reserved-categories', methods=['GET'])
def get_reserved_categories():
    recs = ReservedCategoryLookup.query.all()
    if not recs:
        defaults = ["SC", "ST", "OBC", "Divyangjan", "Gen-EWS", "Others"]
        for d in defaults: db.session.add(ReservedCategoryLookup(value=d))
        db.session.commit()
        recs = ReservedCategoryLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/library-resources', methods=['GET'])
def get_library_resources():
    recs = LibraryResourceLookup.query.all()
    if not recs:
        defaults = ["Books", "Journals", "e-Journals", "e-Books", "e-ShodhSindhu", "Shodhganga", "Databases"]
        for d in defaults: db.session.add(LibraryResourceLookup(value=d))
        db.session.commit()
        recs = LibraryResourceLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/qualifying-exams', methods=['GET'])
def get_qualifying_exams():
    recs = QualifyingExamLookup.query.all()
    if not recs:
        defaults = ["NET", "SLET", "GATE", "GMAT", "CAT", "GRE", "JAM", "IELTS", "TOEFL", "Civil Services", "State government examinations", "Other examinations"]
        for d in defaults: db.session.add(QualifyingExamLookup(value=d))
        db.session.commit()
        recs = QualifyingExamLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/event-levels', methods=['GET'])
def get_event_levels():
    recs = EventLevelLookup.query.all()
    if not recs:
        defaults = ["Inter-university", "State", "National", "International", "District"]
        for d in defaults: db.session.add(EventLevelLookup(value=d))
        db.session.commit()
        recs = EventLevelLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/award-categories', methods=['GET'])
def get_award_categories():
    recs = AwardCategoryLookup.query.all()
    if not recs:
        defaults = ["Individual", "Team"]
        for d in defaults: db.session.add(AwardCategoryLookup(value=d))
        db.session.commit()
        recs = AwardCategoryLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

# --- Generic POST Endpoint to add new lookup items ---
LOOKUP_MODELS = {
    "semesters": SemesterLookup,
    "course-types": CourseTypeLookup,
    "academic-years": AcademicYearLookup,
    "reserved-categories": ReservedCategoryLookup,
    "library-resources": LibraryResourceLookup,
    "qualifying-exams": QualifyingExamLookup,
    "event-levels": EventLevelLookup,
    "award-categories": AwardCategoryLookup,
    "departments": DepartmentLookup,
    "designations": DesignationLookup,
    "highest-degrees": HighestDegreeLookup,
    "appointment-types": AppointmentTypeLookup,
    "funding-agencies": FundingAgencyLookup,
    "levels": LevelLookup,
    "team-individual": TeamIndividualLookup,
    "egovernance-areas": EGovernanceAreaLookup
}

# --- Dedicated Getters for new Lookups ---

@api_bp.route('/get-lookups/<lookup_key>', methods=['GET'])
def get_lookup_values(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model: return jsonify([])
    recs = model.query.all()
    # Handle different models if needed, but most use .value
    return jsonify([{"id": r.id, "value": r.value, "label": r.value} for r in recs])

@api_bp.route('/lookups/<lookup_key>', methods=['POST'])
def add_lookup_value(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model:
        return jsonify({"success": False, "error": "Invalid lookup key"}), 404
        
    data = request.json
    new_val = data.get('value')
    if not new_val:
        return jsonify({"success": False, "error": "Value required"}), 400
        
    existing = model.query.filter_by(value=new_val).first()
    if not existing:
        try:
            new_item = model(value=new_val)
            db.session.add(new_item)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 400
            
    return jsonify({"success": True, "message": f"Added {new_val}"})

# --- Generic Criteria Router ---

CRITERIA_MODELS = {
    "1_1": C11Courses, "1_1_3": C113TeacherBodies, "1_2_1": C121CBCS, "1_2_2": C122Addon, 
    "1_3_2": C132Experiential, "1_3_3": C133Projects, 
    "2_1_1": C211Enrolment, "2_1_2": C212Reservation, "2_3": C23OutgoingStudents,
    "2_3_3": C233MentorRatio, "2_4_1": C241Teachers, "2_4_2": C242TeacherPhD, "2_6_3": C263PassPercentage,
    # Criteria 3
    "3_1": C3FullTimeTeachers, "3_2": C3SanctionedPosts,
    "3_1_1_2": C3ResearchProjects, "3_1_3": C313Events, "3_2_1": C321Papers,
    "3_2_2": C322Books, "3_3_2": C332ExtensionAwards, "3_3_3_4": C333Outreach,
    "3_4_1": C341Collaborations, "3_4_2": C342MoUs,
    # Criteria 4
    "4_1_3": C413ICTRooms, "4_1_4": C4Expenditure, "4_2_2": C42Library,
    "5_1_1": C511Scholarships, "5_1_3": C513SkillInitiatives, "5_1_4": C514CompetitiveExams,
    "5_2_1": C521Placements, "5_2_2": C522HigherEd, "5_2_3": C523QualifyingExams,
    "5_3_1": C531SportsAwards, "5_3_3": C533SportsEvents,
    # Criteria 6
    "6_2_3": C623EGovernance, "6_3_2": C632TeacherFinancial, "6_3_3": C633StaffTraining,
    "6_3_4": C634TeacherFDP, "6_5_3": C653QualityInitiatives,
}

from decimal import Decimal
from datetime import date

def to_dict(rec):
    d = {}
    for col in rec.__table__.columns:
        val = getattr(rec, col.name)
        if isinstance(val, date): val = val.isoformat()
        if isinstance(val, Decimal): val = float(val)
        if isinstance(val, datetime): val = val.isoformat()
        d[col.name] = val
        
        # UI Aliasing
        if col.name == 'academic_year' or col.name == 'year_of_offering':
            d['year'] = val
        if col.name == 'status_of_implementation':
            d['cbcsStatus'] = "Yes" if val else "No"
        if col.name == 'year_of_implementation':
            d['cbcsYear'] = val
            
        DB_TO_UI = {
            'room_number': 'room_details', 'ict_facility_type': 'ict_type', 'geo_tagged_photo_link': 'link',
            'resource_type': 'resource', 'membership_details': 'details',
            'expenditure_ejournals': 'exp_subscription', 'expenditure_eresources': 'exp_others',
            'total_library_expenditure': 'total_exp', 'proof_links': 'link',
            'budget_infra': 'budget_allocation', 'expenditure_infra': 'expenditure_augmentation',
            'total_expenditure_excl_salary': 'total_exp_excluding_salary',
            'expenditure_academic_maint': 'maintenance_academic', 'expenditure_physical_maint': 'maintenance_physical',
            'co_investigator': 'pi_name', 'activity_report_link': 'date_from_to',
            'supporting_document': 'upload_supporting_document'
        }
        if col.name in DB_TO_UI:
            d[DB_TO_UI[col.name]] = val


    # Lookup related names for the UI tables
    if hasattr(rec, 'course_id') and rec.course_id:
        c = Course.query.get(rec.course_id)
        if c: d['courseCode'] = c.course_code; d['courseName'] = c.course_name
    if hasattr(rec, 'program_id') and rec.program_id:
        p = Program.query.get(rec.program_id)
        if p: d['programCode'] = p.program_code; d['programName'] = p.program_name
    if hasattr(rec, 'teacher_id') and rec.teacher_id:
        t = Teacher.query.get(rec.teacher_id)
        if t: d['teacherName'] = t.name
    if hasattr(rec, 'student_id') and rec.student_id:
        s = Student.query.get(rec.student_id)
        if s: d['enrollmentNumber'] = s.enrollment_number; d['studentName'] = s.name
            
    # Resolve Usernames for auditing
    if hasattr(rec, 'created_by_id') and rec.created_by_id:
        u = User.query.get(rec.created_by_id)
        if u: d['createdBy'] = u.username
    if hasattr(rec, 'updated_by_id') and rec.updated_by_id:
        u = User.query.get(rec.updated_by_id)
        if u: d['updatedBy'] = u.username

    return d

@api_bp.route('/records/<criterion>', methods=['GET'])
def get_records(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify([])
    
    if criterion == '1_1':
        # Enrich with semester name (stored as programName in C11Courses via program relation)
        records = C11Courses.query.all()
        result = []
        for r in records:
            d = to_dict(r)
            # Also resolve semester from SemesterCourseLookup using course_code
            if r.course_id:
                c = Course.query.get(r.course_id)
                if c:
                    scl = SemesterCourseLookup.query.filter_by(course_code=c.course_code).first()
                    if scl: d['programName'] = scl.semester
            result.append(d)
        return jsonify(result)
    
    # ---- Enrich 3_2_2: add teacher names for display ----
    if criterion == '3_2_2':
        records = C322Books.query.all()
        result = []
        for r in records:
            d = to_dict(r)
            if r.teacher_id:
                t = Teacher.query.get(r.teacher_id)
                d['author_names'] = t.name if t else ''
                d['teacher_ids'] = [str(r.teacher_id)]
            else:
                d['author_names'] = ''
                d['teacher_ids'] = []
            result.append(d)
        return jsonify(result)

    # ---- Enrich 3_1_1_2: remap co_investigator → pi_name for UI ----
    if criterion == '3_1_1_2':
        records = C3ResearchProjects.query.all()
        result = []
        for r in records:
            d = to_dict(r)
            d['pi_name'] = r.co_investigator or ''  # alias for UI
            result.append(d)
        return jsonify(result)

    return jsonify([to_dict(r) for r in model.query.all()])

# ---- Dedicated POST for Criterion 1.1 ----
@api_bp.route('/records/1_1', methods=['POST'])
def add_record_1_1():
    """
    Handles Criterion 1.1 form submission.
    UI sends: { department, programCode, programName (=semester), courseCode, courseName, year }
    Maps to C11Courses: { academic_year, program_id, course_id }
    """
    data = request.json or {}
    
    semester    = data.get('programName', '').strip()  # "FYMCA-SEM-I", etc.
    program_code= data.get('programCode', '').strip()
    course_code = data.get('courseCode', '').strip()
    year        = str(data.get('year', '')).strip()
    department  = data.get('department', '').strip()

    # --- Validate required fields ---
    if not semester:
        return jsonify({"success": False, "error": "Semester (Program Name) is required"}), 400
    if not course_code:
        return jsonify({"success": False, "error": "Course Code is required"}), 400
    if not year:
        return jsonify({"success": False, "error": "Year is required"}), 400

    # --- 1. Resolve / auto-create Program ---
    prog = Program.query.filter_by(program_code=program_code).first() if program_code else None
    if not prog:
        # Create a minimal program entry so FK is satisfied
        prog = Program(
            program_code = program_code or semester,
            program_name = semester,
            department   = department
        )
        db.session.add(prog)
        db.session.flush()   # get prog.id without committing

    # --- 2. Resolve / auto-create Course (in the Course dim table) ---
    course = Course.query.filter_by(course_code=course_code).first()
    if not course:
        # Pull course name from the lookup table
        scl = SemesterCourseLookup.query.filter_by(
            semester=semester, course_code=course_code
        ).first()
        course_name = scl.course_name if scl else data.get('courseName', course_code)
        course = Course(course_code=course_code, course_name=course_name)
        db.session.add(course)
        db.session.flush()

    # --- 3. Duplicate check: same year + program + course ---
    existing = C11Courses.query.filter_by(
        academic_year=year,
        program_id=prog.id,
        course_id=course.id
    ).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {course_code} for {semester} in {year} already exists"}), 400

    # --- 4. Insert record ---
    try:
        new_rec = C11Courses(
            academic_year = year,
            program_id    = prog.id,
            course_id     = course.id,
            proof_links   = data.get('proofLink'),
            created_by_id = session.get('user_id'),
            updated_by_id = session.get('user_id')
        )
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

# ---- Dedicated POST for Criterion 3.1.1 & 3.1.2 (Research Projects) ----
@api_bp.route('/records/3_1_1_2', methods=['POST'])
def add_record_3_1_1_2():
    """
    Maps: project_name, pi_name (→ co_investigator), department,
          year_of_award, amount_sanctioned, duration, funding_agency, funding_type
    """
    data = request.json or {}
    try:
        rec = C3ResearchProjects(
            project_name      = data.get('project_name', ''),
            co_investigator   = data.get('pi_name', data.get('co_investigator', '')),
            department        = data.get('department', ''),
            year_of_award     = int(data['year_of_award']) if data.get('year_of_award') and str(data['year_of_award']).isdigit() else None,
            amount_sanctioned = data.get('amount_sanctioned') or None,
            duration          = data.get('duration', ''),
            funding_agency    = data.get('funding_agency', ''),
            funding_type      = data.get('funding_type', ''),
            created_by_id     = session.get('user_id'),
            updated_by_id     = session.get('user_id'),
        )
        db.session.add(rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

# ---- Dedicated POST for Criterion 3.2.2 (Books/Chapters per Teacher) ----
@api_bp.route('/records/3_2_2', methods=['POST'])
def add_record_3_2_2():
    """
    Frontend sends teacher_ids (array of IDs) + other_teacher_name.
    We store the primary teacher_id (first in list) and put full names in author_names.
    """
    data = request.json or {}
    teacher_ids  = data.get('teacher_ids', [])
    other_name   = data.get('other_teacher_name', '')

    # Resolve teacher names for display / audit
    names = []
    primary_tid = None
    for tid in teacher_ids:
        t = Teacher.query.get(int(tid))
        if t:
            names.append(t.name)
            if primary_tid is None:
                primary_tid = t.id
    if other_name:
        names.append(other_name)

    try:
        rec = C322Books(
            teacher_id           = primary_tid,
            book_title           = data.get('book_title', ''),
            paper_title          = data.get('paper_title', ''),
            chapter_title        = data.get('book_title', ''),   # alias
            conference_name      = data.get('conference_name', ''),
            proceedings_title    = data.get('proceedings_title', ''),
            level                = data.get('level', ''),
            year_of_publication  = data.get('year_of_publication') or None,
            isbn_issn            = data.get('isbn_issn', ''),
            affiliating_institute= data.get('affiliating_institute', ''),
            publisher            = data.get('publisher', ''),
            created_by_id        = session.get('user_id'),
            updated_by_id        = session.get('user_id'),
        )
        db.session.add(rec)
        db.session.commit()
        d = to_dict(rec)
        d['author_names'] = ", ".join(names)  # add for UI display
        d['teacher_ids']  = teacher_ids
        return jsonify({"success": True, "data": d})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>', methods=['POST'])
def add_record(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False, "error": "Criterion not found"}), 404
    
    data = request.json
    db_kwargs = {}
    
    # Audit tracking
    user_id = session.get('user_id')
    db_kwargs['created_by_id'] = user_id
    db_kwargs['updated_by_id'] = user_id
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    # 1. Map common foreign keys dynamically based on UI selections
    if 'courseCode' in data:
        c = Course.query.filter_by(course_code=data['courseCode']).first()
        if c and hasattr(model, 'course_id'): db_kwargs['course_id'] = c.id
    if 'programCode' in data:
        p = Program.query.filter_by(program_code=data['programCode']).first()
        if p and hasattr(model, 'program_id'): db_kwargs['program_id'] = p.id
    if 'teacherName' in data:
        t = Teacher.query.filter_by(name=data['teacherName']).first()
        if t and hasattr(model, 'teacher_id'): db_kwargs['teacher_id'] = t.id
        
    # 2. Map payload keys to model columns
    # UI field name -> DB column name aliases
    UI_TO_DB = {
        'room_details': 'room_number', 'ict_type': 'ict_facility_type', 'link': 'geo_tagged_photo_link',
        'resource': 'resource_type', 'details': 'membership_details',
        'exp_subscription': 'expenditure_ejournals', 'exp_others': 'expenditure_eresources',
        'total_exp': 'total_library_expenditure', 'link': 'proof_links',
        'budget_allocation': 'budget_infra', 'expenditure_augmentation': 'expenditure_infra',
        'total_exp_excluding_salary': 'total_expenditure_excl_salary',
        'maintenance_academic': 'expenditure_academic_maint', 'maintenance_physical': 'expenditure_physical_maint',
        'pi_name': 'co_investigator', 'date_from_to': 'activity_report_link',
        'upload_supporting_document': 'supporting_document'
    }
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        
        # Direct match
        if col_name in data and data[col_name] is not None:
            if col_name == 'status_of_implementation':
                 db_kwargs[col_name] = str(data[col_name]).lower() == "yes"
            else:
                 db_kwargs[col_name] = data[col_name]
        else:
            # Check UI alias
            found_ui = False
            for ui_key, db_key in UI_TO_DB.items():
                if db_key == col_name and ui_key in data:
                    db_kwargs[col_name] = data[ui_key]
                    found_ui = True
                    break
            if found_ui:
                continue
            # Common UI -> DB Alias mapping
            if col_name == 'academic_year' and 'year' in data:
                db_kwargs[col_name] = data['year']
            elif col_name == 'year_of_offering' and 'year' in data:
                db_kwargs[col_name] = int(data['year']) if str(data['year']).isdigit() else None
            elif col_name == 'year_of_passing' and 'year' in data:
                db_kwargs[col_name] = int(data['year']) if str(data['year']).isdigit() else None
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                db_kwargs[col_name] = int(data['cbcsYear']) if str(data['cbcsYear']).isdigit() else None
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                db_kwargs[col_name] = str(data['cbcsStatus']).lower() == "yes"

    # Default Academic Year from session if missing in payload
    if 'academic_year' not in db_kwargs and hasattr(model, 'academic_year'):
        db_kwargs['academic_year'] = session.get('academic_year')

    try:
        new_rec = model(**db_kwargs)
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>/<int:id>', methods=['PUT'])
def update_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False}), 404
    rec = model.query.get(id)
    if not rec: return jsonify({"success": False}), 404
    
    data = request.json
    
    # Audit tracking
    rec.updated_by_id = session.get('user_id')
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    if 'courseCode' in data:
        c = Course.query.filter_by(course_code=data['courseCode']).first()
        if c and hasattr(model, 'course_id'): rec.course_id = c.id
    if 'programCode' in data:
        p = Program.query.filter_by(program_code=data['programCode']).first()
        if p and hasattr(model, 'program_id'): rec.program_id = p.id
        
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        if col_name in data:
            if col_name == 'status_of_implementation':
                setattr(rec, col_name, str(data[col_name]).lower() == "yes")
            else:
                setattr(rec, col_name, data[col_name])
        else:
            if col_name == 'academic_year' and 'year' in data:
                setattr(rec, col_name, data['year'])
            elif col_name == 'year_of_offering' and 'year' in data:
                setattr(rec, col_name, int(data['year']) if str(data['year']).isdigit() else getattr(rec, col_name))
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                setattr(rec, col_name, int(data['cbcsYear']) if str(data['cbcsYear']).isdigit() else getattr(rec, col_name))
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                setattr(rec, col_name, str(data['cbcsStatus']).lower() == "yes")
    db.session.commit()
    return jsonify({"success": True, "data": to_dict(rec)})

@api_bp.route('/records/<criterion>/<int:id>', methods=['DELETE'])
def delete_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        rec = model.query.get(id)
        if rec:
            db.session.delete(rec)
            db.session.commit()
    return jsonify({"success": True})

@api_bp.route('/records/<criterion>/bulk-delete', methods=['POST'])
def bulk_delete(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        ids = request.json.get('ids', [])
        model.query.filter(model.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
    return jsonify({"success": True})


# ============================================================
# EXCEL EXPORT ENDPOINTS — /api/export-excel/<criterion>
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def _make_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def _hdr(ws, row, col, val, bg='1F4E79'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _make_border()
    return c

def _dat(ws, row, col, val, center=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Arial', size=10)
    c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    c.border = _make_border()
    return c

def _title(ws, title, cols):
    ws.merge_cells(f'A1:{get_column_letter(cols)}1')
    c = ws['A1']
    c.value = title
    c.font = Font(bold=True, size=12, color='FFFFFF', name='Arial')
    c.fill = PatternFill('solid', start_color='C00000')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _make_border()
    ws.row_dimensions[1].height = 28

def _send_wb(wb, filename):
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@api_bp.route('/export-excel/<criterion>')
def export_excel(criterion):
    """Generic + specific Excel export for all criteria."""
    # ---- Criteria 3 ----
    if criterion == '3_1':
        records = C3FullTimeTeachers.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.1'
        _title(ws, '3.1 Number of Full-Time Teachers During the Year', 4)
        hdrs = ['#','Teacher Name','Academic Year','Sanctioned Posts']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            t = Teacher.query.get(rec.teacher_id); tname = t.name if t else ''
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,tname); _dat(ws,r,3,rec.academic_year,True); _dat(ws,r,4,rec.sanctioned_posts,True)
        for i,w in enumerate([4,35,15,18],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_1_FullTimeTeachers.xlsx')

    if criterion == '3_2':
        records = C3SanctionedPosts.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.2'
        _title(ws, '3.2 Number of Sanctioned Posts During the Year', 4)
        hdrs = ['#','Year','Number of Sanctioned Posts','Proof Link']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.year,True)
            _dat(ws,r,3,rec.sanctioned_posts_count,True); _dat(ws,r,4,rec.supporting_document)
        for i,w in enumerate([4,15,25,45],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_2_SanctionedPosts.xlsx')

    if criterion == '3_1_1_2':
        records = C3ResearchProjects.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.1.1-3.1.2'
        _title(ws, '3.1.1 & 3.1.2 Research Grants & Departments with Research Projects', 8)
        hdrs = ['#','Project / Endowments','Principal Investigator','Department','Year of Award','Amount Sanctioned (₹L)','Duration','Funding Agency','Type']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.project_name); _dat(ws,r,3,rec.co_investigator)
            _dat(ws,r,4,rec.department); _dat(ws,r,5,rec.year_of_award,True)
            _dat(ws,r,6,float(rec.amount_sanctioned) if rec.amount_sanctioned else '',True)
            _dat(ws,r,7,rec.duration); _dat(ws,r,8,rec.funding_agency); _dat(ws,r,9,rec.funding_type)
        for i,w in enumerate([4,40,25,20,12,18,15,25,15],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_1_Research_Grants.xlsx')

    if criterion == '3_1_3':
        records = C313Events.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.1.3'
        _title(ws, '3.1.3 Seminars / Conferences / Workshops Conducted', 5)
        hdrs = ['#','Year','Name of Workshop/Seminar','Participants','Date From','Date To','Activity Report Link']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.year,True); _dat(ws,r,3,rec.event_name)
            _dat(ws,r,4,rec.participant_count,True)
            _dat(ws,r,5,rec.date_from.isoformat() if rec.date_from else '',True)
            _dat(ws,r,6,rec.date_to.isoformat() if rec.date_to else '',True)
            _dat(ws,r,7,rec.activity_report_link)
        for i,w in enumerate([4,12,40,14,14,14,45],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_1_3_Events.xlsx')

    if criterion == '3_2_1':
        records = C321Papers.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.2.1'
        _title(ws, '3.2.1 Papers Published per Teacher in UGC Notified Journals', 7)
        hdrs = ['#','Title of Paper','Author/s','Department','Journal Name','Year','ISSN','UGC Proof Link']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.paper_title); _dat(ws,r,3,rec.author_names)
            _dat(ws,r,4,rec.department); _dat(ws,r,5,rec.journal_name)
            _dat(ws,r,6,rec.year_of_publication,True); _dat(ws,r,7,rec.issn,True)
            _dat(ws,r,8,rec.ugc_recognition_link)
        for i,w in enumerate([4,40,25,20,30,12,16,45],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_2_1_Papers.xlsx')

    if criterion == '3_2_2':
        records = C322Books.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.2.2'
        _title(ws, '3.2.2 Books, Chapters & Conference Proceedings per Teacher', 10)
        hdrs = ['#','Teacher','Book/Chapter Title','Paper Title','Proceedings Title','Conference','Level','Year','ISBN/ISSN','Affiliating Institute','Publisher']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            t = Teacher.query.get(rec.teacher_id); tname = t.name if t else ''
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,tname); _dat(ws,r,3,rec.book_title)
            _dat(ws,r,4,rec.paper_title); _dat(ws,r,5,rec.proceedings_title)
            _dat(ws,r,6,rec.conference_name); _dat(ws,r,7,rec.level,True)
            _dat(ws,r,8,rec.year_of_publication,True); _dat(ws,r,9,rec.isbn_issn,True)
            _dat(ws,r,10,rec.affiliating_institute); _dat(ws,r,11,rec.publisher)
        for i,w in enumerate([4,25,35,30,30,25,14,12,16,30,25],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_2_2_Books.xlsx')

    if criterion == '3_3_2':
        records = C332ExtensionAwards.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.3.2'
        _title(ws, '3.3.2 Extension & Outreach Awards', 4)
        hdrs = ['#','Activity Name','Award Name','Awarding Body','Year']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.activity_name); _dat(ws,r,3,rec.award_name)
            _dat(ws,r,4,rec.awarding_body); _dat(ws,r,5,rec.award_year,True)
        for i,w in enumerate([4,35,30,30,12],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_3_2_Awards.xlsx')

    if criterion == '3_3_3_4':
        records = C333Outreach.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.3.3-4'
        _title(ws, '3.3.3 & 3.3.4 Outreach & Extension Activities', 5)
        hdrs = ['#','Activity Name','Agency Name','Scheme Name','Year','Students Participated']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.activity_name); _dat(ws,r,3,rec.agency_name)
            _dat(ws,r,4,rec.scheme_name); _dat(ws,r,5,rec.year,True); _dat(ws,r,6,rec.students_participated,True)
        for i,w in enumerate([4,35,30,25,12,18],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_3_3_4_Outreach.xlsx')

    if criterion == '3_4_1':
        records = C341Collaborations.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.4.1'
        _title(ws, '3.4.1 Linkages / Functional Collaborations', 6)
        hdrs = ['#','Activity Title','Agency Name','Participant','Year','Duration','Nature of Activity','Proof Link']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.activity_title); _dat(ws,r,3,rec.agency_name)
            _dat(ws,r,4,rec.participant_name); _dat(ws,r,5,rec.year,True)
            _dat(ws,r,6,rec.duration); _dat(ws,r,7,rec.nature_of_activity); _dat(ws,r,8,rec.proof_links)
        for i,w in enumerate([4,30,25,25,12,15,25,40],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_4_1_Collaborations.xlsx')

    if criterion == '3_4_2':
        records = C342MoUs.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C3.4.2'
        _title(ws, '3.4.2 MoUs / Functional Linkages', 6)
        hdrs = ['#','Organisation Name','Institution Name','Signing Year','Duration','Activities','Participants']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.organisation_name); _dat(ws,r,3,rec.institution_name)
            _dat(ws,r,4,rec.signing_year,True); _dat(ws,r,5,rec.duration)
            _dat(ws,r,6,rec.activities_list); _dat(ws,r,7,rec.participant_count,True)
        for i,w in enumerate([4,30,30,14,15,35,14],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_3_4_2_MoUs.xlsx')

    # ---- Criteria 4 ----
    if criterion == '4_1_3':
        records = C413ICTRooms.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C4.1.3'
        _title(ws, '4.1.3 ICT-Enabled Classrooms / Smart Rooms', 3)
        hdrs = ['#','Room / Classroom / Seminar Hall','Type of ICT Facility','Link to Geo-Tagged Photos']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.room_number); _dat(ws,r,3,rec.ict_facility_type)
            _dat(ws,r,4,rec.geo_tagged_photo_link)
        for i,w in enumerate([4,35,30,45],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_4_1_3_ICT.xlsx')

    if criterion == '4_1_4':
        records = C4Expenditure.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C4.1.4'
        _title(ws, '4.1.4 & 4.4.1 Infrastructure Expenditure (INR in Lakhs)', 6)
        hdrs = ['#','Year','Budget Allocated (₹L)','Expenditure Augmentation (₹L)','Total Exp. Excl. Salary (₹L)','Maint. Academic (₹L)','Maint. Physical (₹L)']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.year,True)
            _dat(ws,r,3,float(rec.budget_infra) if rec.budget_infra else '',True)
            _dat(ws,r,4,float(rec.expenditure_infra) if rec.expenditure_infra else '',True)
            _dat(ws,r,5,float(rec.total_expenditure_excl_salary) if rec.total_expenditure_excl_salary else '',True)
            _dat(ws,r,6,float(rec.expenditure_academic_maint) if rec.expenditure_academic_maint else '',True)
            _dat(ws,r,7,float(rec.expenditure_physical_maint) if rec.expenditure_physical_maint else '',True)
        for i,w in enumerate([4,14,20,22,22,20,20],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_4_1_4_Expenditure.xlsx')

    if criterion == '4_2_2':
        records = C42Library.query.all()
        wb = Workbook(); ws = wb.active; ws.title = 'C4.2.2'
        _title(ws, '4.2.2 & 4.2.3 Library Resources & Expenditure', 6)
        hdrs = ['#','Academic Year','Resource Type','Membership Details','Exp. e-Journals (₹L)','Exp. Others (₹L)','Total Library Exp. (₹L)','Proof Link']
        for i,h in enumerate(hdrs,1): _hdr(ws,2,i,h)
        for r,rec in enumerate(records,3):
            _dat(ws,r,1,r-2,True); _dat(ws,r,2,rec.academic_year,True)
            _dat(ws,r,3,rec.resource_type); _dat(ws,r,4,rec.membership_details)
            _dat(ws,r,5,float(rec.expenditure_ejournals) if rec.expenditure_ejournals else '',True)
            _dat(ws,r,6,float(rec.expenditure_eresources) if rec.expenditure_eresources else '',True)
            _dat(ws,r,7,float(rec.total_library_expenditure) if rec.total_library_expenditure else '',True)
            _dat(ws,r,8,rec.proof_links)
        for i,w in enumerate([4,16,18,30,18,14,18,40],1): ws.column_dimensions[get_column_letter(i)].width=w
        return _send_wb(wb,'Criteria_4_2_2_Library.xlsx')

    # ---- Generic fallback ----
    model = CRITERIA_MODELS.get(criterion)
    if not model:
        return jsonify({"error": "Unknown criterion"}), 404
    records = model.query.all()
    wb = Workbook(); ws = wb.active; ws.title = criterion
    cols = [c.name for c in model.__table__.columns]
    for i,h in enumerate(cols,1): _hdr(ws,1,i,h)
    for r,rec in enumerate(records,2):
        for i,col in enumerate(cols,1):
            val = getattr(rec,col)
            if isinstance(val, date): val = val.isoformat()
            if isinstance(val, Decimal): val = float(val)
            _dat(ws,r,i,val)
    return _send_wb(wb, f'Criteria_{criterion}.xlsx')
